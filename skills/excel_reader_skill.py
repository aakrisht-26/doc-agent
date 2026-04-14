"""
ExcelReaderSkill — extracts structured text and data from Excel and CSV files.

Supported formats:
    .xlsx / .xls   — via openpyxl (structure-aware, handles merged cells)
    .csv           — via pandas (encoding auto-detection)

Each sheet becomes one DocumentChunk. Tables are preserved as text representations
for downstream summarization and question extraction.
"""

from __future__ import annotations

import time
from pathlib import Path
from typing import Any, Dict, List, Optional

from core.models import DocumentChunk, ParsedDocument, SkillInput, SkillOutput
from skills.base_skill import BaseSkill
from utils.logger import get_logger

logger = get_logger(__name__)


class ExcelReaderSkill(BaseSkill):
    """
    Reads and parses Excel / CSV files into a structured ParsedDocument.

    Config keys:
        max_sheets        (int)  : Max number of sheets to process (default: 50)
        include_formulas  (bool) : Include formula text instead of values (default: False)
        max_rows_per_sheet(int)  : Row limit per sheet (default: 10000)
    """

    name = "excel_reader"
    description = "Extracts data, headers, and text from Excel / CSV files, per sheet."
    required_inputs = ["file_path"]

    def __init__(self, config: Optional[Dict[str, Any]] = None) -> None:
        super().__init__(config)
        self._max_sheets   = self.get_config("max_sheets", 50)
        self._max_rows     = self.get_config("max_rows_per_sheet", 10_000)
        self._incl_formula = self.get_config("include_formulas", False)

    # ── Skill entry point ─────────────────────────────────────────────

    def execute(self, inputs: SkillInput) -> SkillOutput:
        start = time.monotonic()
        file_path = Path(inputs.data["file_path"])

        if not file_path.exists():
            return SkillOutput(success=False, data=None,
                               error=f"File not found: {file_path}")

        suffix = file_path.suffix.lower()
        try:
            if suffix == ".csv":
                doc = self._parse_csv(file_path)
            else:
                doc = self._parse_excel(file_path)
        except Exception as exc:
            self.logger.error(f"Excel read failed: {exc}", exc_info=True)
            return SkillOutput(
                success=False, data=None,
                error=f"Failed to parse '{file_path.name}': {exc}",
                duration_ms=(time.monotonic() - start) * 1000,
            )

        warnings: List[str] = []
        if doc.is_empty:
            warnings.append("No data extracted. The file may be empty or all sheets are blank.")

        self.logger.info(
            f"Excel parsed: {doc.page_count} sheet(s), "
            f"{doc.word_count} words, {len(doc.tables)} table(s)"
        )
        return SkillOutput(
            success=True,
            data=doc,
            warnings=warnings,
            duration_ms=(time.monotonic() - start) * 1000,
        )

    # ── Parsers ───────────────────────────────────────────────────────

    def _parse_excel(self, file_path: Path) -> ParsedDocument:
        """Parse .xlsx / .xls files using openpyxl."""
        import openpyxl
        import pandas as pd

        data_only = not self._incl_formula
        wb = openpyxl.load_workbook(str(file_path), data_only=data_only)

        chunks: List[DocumentChunk] = []
        tables: List[Dict[str, Any]] = []
        text_parts: List[str] = []
        sheet_names: List[str] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames[: self._max_sheets]):
            ws = wb[sheet_name]
            sheet_names.append(sheet_name)

            # Collect non-empty rows
            rows: List[List[str]] = []
            for row in ws.iter_rows(max_row=self._max_rows, values_only=True):
                if any(cell is not None for cell in row):
                    rows.append([str(c).strip() if c is not None else "" for c in row])

            if not rows:
                self.logger.debug(f"Sheet '{sheet_name}' is empty — skipped.")
                continue

            df = self._rows_to_dataframe(rows)
            text_repr = self._dataframe_to_text(df, sheet_name)
            text_parts.append(text_repr)

            tables.append({
                "sheet": sheet_name,
                "sheet_index": sheet_idx,
                "rows": len(df),
                "columns": len(df.columns),
                "preview": rows[:5],
            })
            chunks.append(DocumentChunk(
                text=text_repr,
                page_or_sheet=sheet_name,
                chunk_index=sheet_idx,
                metadata={"sheet": sheet_name, "rows": len(rows)},
            ))

        wb.close()

        return ParsedDocument(
            file_name=file_path.name,
            file_type="excel",
            chunks=chunks,
            full_text="\n\n" + ("─" * 60) + "\n\n".join(text_parts),
            tables=tables,
            metadata={"sheets": len(sheet_names), "sheet_names": sheet_names},
            page_count=len(sheet_names) if sheet_names else 0,
            sheet_names=sheet_names,
        )

    def _parse_csv(self, file_path: Path) -> ParsedDocument:
        """Parse .csv files using pandas with encoding fallback."""
        import pandas as pd

        for enc in ("utf-8", "latin-1", "cp1252"):
            try:
                df = pd.read_csv(str(file_path), nrows=self._max_rows, encoding=enc)
                break
            except (UnicodeDecodeError, Exception):
                continue
        else:
            raise ValueError("Could not decode CSV with any known encoding.")

        text_repr = self._dataframe_to_text(df, file_path.stem)
        return ParsedDocument(
            file_name=file_path.name,
            file_type="excel",
            chunks=[DocumentChunk(
                text=text_repr,
                page_or_sheet=1,
                chunk_index=0,
                metadata={"source": "csv", "rows": len(df), "cols": len(df.columns)},
            )],
            full_text=text_repr,
            tables=[{"sheet": "CSV", "rows": len(df), "columns": len(df.columns)}],
            metadata={"format": "csv", "rows": len(df), "columns": list(df.columns)},
            page_count=1,
            sheet_names=["CSV"],
        )

    # ── Helpers ───────────────────────────────────────────────────────

    @staticmethod
    def _rows_to_dataframe(rows: List[List[str]]):
        """Convert list-of-rows to a DataFrame, promoting first row to header if suitable."""
        import pandas as pd

        df = pd.DataFrame(rows)
        if rows and ExcelReaderSkill._looks_like_header(rows[0]):
            df.columns = [str(h).strip() or f"col_{i}" for i, h in enumerate(rows[0])]
            df = df.iloc[1:].reset_index(drop=True)
        return df

    @staticmethod
    def _looks_like_header(row: List[str]) -> bool:
        """Heuristic: header rows have mostly non-numeric text."""
        non_empty = [c for c in row if c.strip()]
        if not non_empty:
            return False
        numeric = sum(
            1 for c in non_empty
            if c.replace(".", "").replace("-", "").replace(",", "").isdigit()
        )
        return numeric / len(non_empty) < 0.5

    @staticmethod
    def _dataframe_to_text(df: pd.DataFrame, sheet_name: str) -> str:
        """Convert DataFrame to a human-readable text block with smart sampling."""
        import pandas as pd

        row_count = len(df)
        lines = [f"[Sheet: {sheet_name}]", f"- Total Rows: {row_count}", ""]
        
        try:
            # If the dataset is large, take head + tail to show breadth
            if row_count > 600:
                head = df.head(300)
                tail = df.tail(300)
                lines.append(head.to_string(index=False, max_cols=30))
                lines.append("\n... [Rows 301 to {row_count-300} omitted] ...\n")
                lines.append(tail.to_string(index=False, max_cols=30))
            else:
                lines.append(df.to_string(index=False, max_cols=30))
        except Exception:
            lines.append(str(df.head(100)))
            
        return "\n".join(lines)
